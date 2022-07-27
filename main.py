import openpyxl as xl
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import DateEntry
from tkinter import filedialog
from datetime import date


def atualizar_tree_contas():
    
    tree_contas.delete(*tree_contas.get_children())
    
    for cell in base_contas['A:A']:
        if cell.row == 1:
            continue
    
        tree_contas.insert('', END, values=(base_contas[f'C{cell.row}'].value, 
                                            base_contas[f'B{cell.row}'].value, 
                                            cell.value))


def limpar_campos_conta():
    
    valor_codigo.delete(0, END)
    valor_descricao_conta.delete(0, END)
    valor_classificacao.delete(0, END)
    valor_classificacao.focus()


def criar_conta():
    if valor_classificacao.get() == '' or valor_descricao_conta.get() == '' or valor_codigo.get() == "":
        messagebox.showerror(title='Erro', message='Preencha todos os campos.')
        return

    Conta(valor_descricao_conta.get(), valor_classificacao.get(), valor_codigo.get(), True)
    

def editar_conta():
    
    if valor_classificacao.get() == '' or valor_descricao_conta.get() == '' or valor_codigo.get() == '':
        messagebox.showerror(title='Erro', message='Preencha todos os campos.')
        return
    
    Conta.edita_conta(int(valor_codigo.get()))


def deletar_conta():
    selecao = tree_contas.focus()
    valores = tree_contas.item(selecao, 'values')
    
    if valores == '':
        messagebox.showerror(title='Erro', message='Selecione uma conta.')
        return
    
    Conta.deleta_conta(int(valores[2]))


def duplo_clique_contas(event):
    limpar_campos_conta()
    selecao = tree_contas.focus()
    valores = tree_contas.item(selecao, 'values')
    
    valor_classificacao.insert(0, valores[0])
    valor_descricao_conta.insert(0, valores[1])
    valor_codigo.insert(0, valores[2])

def atualizar_tree_lancamentos():
    
    tree_lancamentos.delete(*tree_lancamentos.get_children())
    
    for cell in base_lancamentos['A:A']:
        if cell.row == 1:
            continue
    
        tree_lancamentos.insert('', END, values=(cell.value,
                                                 base_lancamentos[f'B{cell.row}'].value, 
                                                 base_lancamentos[f'C{cell.row}'].value, 
                                                 base_lancamentos[f'D{cell.row}'].value,
                                                 base_lancamentos[f'E{cell.row}'].value,
                                                 base_lancamentos[f'F{cell.row}'].value))

def limpar_campos_lancamentos():
    
    valor_data.delete(0, END)
    valor_debito.delete(0, END)
    valor_credito.delete(0, END)
    valor_valor.delete(0, END)
    valor_descricao_lanc.delete('1.0', END)
    valor_data.focus()

def lancar():
    if valor_valor.get() == '' or valor_data.get() == '' or valor_debito.get() == '' or valor_credito.get() == '' or valor_descricao_lanc.get('1.0', 'end-1c') == '':
        messagebox.showerror(title='Erro', message='Preencha todos os campos.')
        return
    
    Lancamento(valor_valor.get(), valor_data.get(), valor_debito.get(), valor_credito.get(), valor_descricao_lanc.get('1.0', 'end-1c'), True)


def editar_lancamento():
    if valor_valor.get() == '' or valor_data.get() == '' or valor_debito.get() == '' or valor_credito.get() == '' or valor_descricao_lanc.get('1.0', 'end-1c') == '':
        messagebox.showerror(title='Erro', message='Preencha todos os campos.')
        print('erro')
        return

    if valor_debito.get() == valor_credito.get():
        messagebox.showerror(title='Erro', message='Conta débito não pode ser igual a conta crédito.')
        return
    
    Lancamento.edita_lancamento(int(valor_id))
    limpar_campos_lancamentos()


def deletar_lancamento():

    selecao = tree_lancamentos.focus()
    valores = tree_lancamentos.item(selecao, 'values')
    
    if valores == '':
        messagebox.showerror(title='Erro', message='Selecione um lançamento.')
        return
    
    Lancamento.deleta_lancamento(int(valores[0]))
    limpar_campos_lancamentos()

def duplo_clique_lancamentos(event):
    limpar_campos_lancamentos()
    
    selecao = tree_lancamentos.focus()
    valores = tree_lancamentos.item(selecao, 'values')
    
    global valor_id
    valor_id = valores[0]
    valor_data.insert(0, valores[1])
    valor_debito.insert(0, valores[2])
    valor_credito.insert(0, valores[3])
    valor_valor.insert(0, valores[4])
    valor_descricao_lanc.insert('1.0', valores[5])


def gerar_razao():
    caminho = filedialog.asksaveasfile(filetypes=[('Excel', '.xlsx')], defaultextension='.xlsx')
    if caminho is None:
        return

    razao_df = xl.Workbook()
    razao = razao_df.active
    razao.append(['Código', 'Data', 'Débito', 'Crédito', 'Valor', 'Descrição'])

    razao_df.save(caminho.name)


def gerar_resumo():
    pass


def gerar_balancete():
    caminho = filedialog.asksaveasfile(filetypes=[('Excel', '.xlsx')], defaultextension='.xlsx')
    if caminho is None:
        return

    balancete_df = xl.Workbook()
    balancete = balancete_df.active
    balancete.append(['Classificação', 'Saldo anterior', 'Débito', 'Crédito', 'Saldo atual'])

    balancete_df.save(caminho.name)


def buscar_conta():
    tree_contas.delete(*tree_contas.get_children())

    for column in base_contas["A:C"]:
        for cell in column: 
            if cell.row == 1:
                continue
            existe = False
            for child in tree_contas.get_children():
                if str(tree_contas.item(child)["values"][2]) == str(base_contas[f'A{cell.row}'].value):
                    existe = True
                    break
            if existe:
                continue                     
            if str(valor_busca_conta.get()) in str(cell.value):
                tree_contas.insert('', END, values=(base_contas[f'C{cell.row}'].value, 
                                                    base_contas[f'B{cell.row}'].value, 
                                                    base_contas[f'A{cell.row}'].value))
    valor_busca_conta.delete(0, END)
        

def buscar_lancamento():
    tree_lancamentos.delete(*tree_lancamentos.get_children())

    for column in base_lancamentos["A:F"]:
        for cell in column: 
            if cell.row == 1:
                continue
            existe = False
            for child in tree_lancamentos.get_children():
                if str(tree_lancamentos.item(child)["values"][2]) == str(base_lancamentos[f'A{cell.row}'].value):
                    existe = True
                    break
            if existe:
                continue                     
            if str(valor_busca_conta.get()) in str(cell.value):
                tree_lancamentos.insert('', END, values=(base_lancamentos[f'A{cell.row}'].value,
                                                         base_lancamentos[f'B{cell.row}'].value, 
                                                         base_lancamentos[f'C{cell.row}'].value, 
                                                         base_lancamentos[f'D{cell.row}'].value,
                                                         base_lancamentos[f'E{cell.row}'].value,
                                                         base_lancamentos[f'F{cell.row}'].value))
    valor_busca_conta.delete(0, END)


class Conta:

    def __init__(self, nome, classificacao, codigo, adicionar_na_base=False):
        
        self.nome = nome
        self.classificacao = classificacao.replace('.', '')
        self.codigo = codigo

        if valor_natureza.get() == "Crédito":
            self.d_ou_c = "C"
        else:
            self.d_ou_c = "D"

        try:
            for cell in base_contas['A']:
                if cell.value == self.codigo:
                    raise ValueError
        except ValueError:
            adicionar_na_base = False

        try:
            self.verifica_classificacao()
        except ValueError:
            adicionar_na_base = False

        linhas = []
        if adicionar_na_base:
            clas = str(self.classificacao)
            invalido = True
            for cell in base_contas['C:C']:

                valor = str(cell.value).replace('.', '')
                linha = cell.row
                if cell.row == 1:
                    continue
                
                if len(clas) == 1:
                
                    if int(valor[0]) > int(clas):
                        self.adiciona_na_base(linha)
                        invalido = False
                        break
                    
                if len(clas) == 2:
                      
                    if valor[0] == clas[0]:
                    
                        linhas.append(cell.row)
                        if clas[1] == 1 or clas[1] == 0:
                            linha = linhas[len(linhas)-1] + 1
                            self.adiciona_na_base(linha)
                            invalido = False
                            break
                        if len(valor) < 2:  
                            continue 
                        if int(valor[1]) > int(clas[1]):
                            self.adiciona_na_base(linha)
                            invalido = False
                            break   

                if len(clas) == 4:

                    if valor[:2] == clas[:2]:

                        print(valor)
                        linhas.append(cell.row)
                        if clas[2:] == '01' or clas[2:] == '00':
                            linha = linhas[len(linhas)-1] + 1
                            self.adiciona_na_base(linha)
                            invalido = False
                            break
                        if len(valor) < 4:
                            continue 
                        if int(valor[2:]) > int(clas[2:]):
                            linha = linhas[len(linhas)-1] + 1
                            self.adiciona_na_base(linha)
                            invalido = False
                            break

                if len(clas) == 6:
                    
                    if valor[:4] == clas[:4]:

                        linhas.append(cell.row)
                        if clas[4:] == '01' or clas[4:] == '00':
                            linha = linhas[len(linhas)-1] + 1
                            self.adiciona_na_base(linha)
                            invalido = False
                            break
                        if len(valor) < 6:  
                            continue 
                        if int(valor[4:]) > int(clas[4:]):
                            self.adiciona_na_base(linha)
                            invalido = False
                            break

                if len(clas) == 10:
                    
                    if valor[:6] == clas[:6]:

                        linhas.append(cell.row)
                        if clas[6:] == '01' or clas[6:] == '00':
                            linha = linhas[len(linhas)-1] + 1
                            self.adiciona_na_base(linha)
                            invalido = False
                            break
                        if len(valor) < 10:  
                            continue 
                        if int(valor[6:]) > int(clas[6:]):
                            self.adiciona_na_base(linha)
                            invalido = False
                            break            
            
            if invalido:
                if len(clas) == 1:
                    self.coloca_ponto()
                    base_contas.append([int(self.codigo), self.nome, self.classificacao, self.d_ou_c, self.a_ou_s])
                if len(clas) == 2:
                    linha = linhas[len(linhas)-1] + 1
                    self.adiciona_na_base(linha)
                if len(clas) == 4:
                    linha = linhas[len(linhas)-1] + 1
                    self.adiciona_na_base(linha)
                if len(clas) == 6:
                    linha = linhas[len(linhas)-1] + 1
                    self.adiciona_na_base(linha)
                if len(clas) == 10:
                    linha = linhas[len(linhas)-1] + 1
                    self.adiciona_na_base(linha)
            df.save('base.xlsx')
            atualizar_tree_contas()
            limpar_campos_conta()
            messagebox.showinfo(title='Sucesso!', message='Conta criada com sucesso.')

    def verifica_classificacao(self):
        
        for cell in base_contas['C']:
            if cell.value.replace('.', '') == self.classificacao:
                messagebox.showerror(title='Erro!', message='Código já existente.')
                raise ValueError
            
            validos = [1, 2, 4, 6, 10]
            if len(self.classificacao) not in validos:                
                msg = 'Classificação inválida.\n\nClassificações válidas:\n\nX\nX.X\nX.X.XX\nX.X.XX.XX\nX.X.XX.XXXX'
                messagebox.showerror(title='Erro!', message=msg)
                raise ValueError
            
    def coloca_ponto(self):
        clas = self.classificacao
        
        if len(clas) == 1:
            self.a_ou_s = 'S'
        if len(clas) == 2:
            self.classificacao = f'{clas[0]}.{clas[1:]}'
            self.a_ou_s = 'S'
        if len(clas) == 4:
            self.classificacao = f'{clas[0]}.{clas[1]}.{clas[2:]}'
            self.a_ou_s = 'S'
        if len(clas) == 6:
            self.classificacao = f'{clas[0]}.{clas[1]}.{clas[2:4]}.{clas[4:]}'
            self.a_ou_s = 'S'
        if len(clas) == 10:
            self.classificacao = f'{clas[0]}.{clas[1]}.{clas[2:4]}.{clas[4:6]}.{clas[6:]}'
            self.a_ou_s = 'A'

    def edita_conta(cod):
        
        for cell in base_contas['A']:
            if cell.row == 1:
                continue
            if int(cell.value) == int(valor_codigo.get()):
                linha = cell.row
                nome = valor_descricao_conta.get()
                classificacao = valor_classificacao.get()
                cod = int(valor_codigo.get())
                #Conta(nome, classificacao, cod)

                base_contas[f'B{linha}'] = nome
                base_contas[f'C{linha}'] = classificacao
                df.save('base.xlsx')
                atualizar_tree_contas()
                limpar_campos_conta()
                messagebox.showinfo(title='Sucesso!', message='Conta alterada com sucesso.')
                break

    def deleta_conta(cod):
        
        for cell in base_contas['A']:
            if cell.row == 1:
                continue
            if int(cell.value) == cod:
                base_contas.delete_rows(cell.row)
                df.save('base.xlsx')
                atualizar_tree_contas()
                limpar_campos_conta()
                break
                #messagebox.showinfo(title='Sucesso!', message='Conta deletada com sucesso.')

    def adiciona_na_base(self, row):
        base_contas.insert_rows(idx=row, amount=1)
        base_contas[f'A{row}'] = int(self.codigo)
        base_contas[f'B{row}'] = self.nome
        self.coloca_ponto()
        base_contas[f'C{row}'] = self.classificacao
        base_contas[f'D{row}'] = self.d_ou_c
        base_contas[f'E{row}'] = self.a_ou_s


class Lancamento:

    global base_lancamentos

    def __init__(self, valor, data, debito, credito, descricao, adicionar_na_base=False):
        
        self.valor = valor
        self.data = data
        self.debito = debito
        self.credito = credito
        self.descricao = descricao
        
        try:
            total_lancamentos = len(base_lancamentos['A'])
            self.cod = base_lancamentos['A'][total_lancamentos - 1].value + 1
        except TypeError:
            self.cod = 1

        try:
            self.verifica_debito()
        except ValueError:
            messagebox.showerror(title='Erro', message='Conta débito inexistente.')
            adicionar_na_base = False

        try:
            self.verifica_credito()
        except ValueError:
            messagebox.showerror(title='Erro', message='Conta crédito inexistente.')
            adicionar_na_base = False
            
        try:    
            if self.credito == self.debito:
                raise ValueError
        except:
            messagebox.showerror(title='Erro', message='Conta débito não pode ser igual a conta crédito.')
            adicionar_na_base = False
        
        if adicionar_na_base:
            base_lancamentos.append([self.cod, self.data, self.debito, self.credito, self.valor, self.descricao])
            df.save('base.xlsx')
            atualizar_tree_lancamentos()
            limpar_campos_lancamentos()
            messagebox.showinfo(title='Sucesso!', message='Lançamento efetuado com sucesso.')

    def verifica_debito(self):
        
        contas = []
        for cell in base_contas['A']:
            contas.append(cell.value)

        if self.debito in contas:
            pass
        else:
            raise ValueError

    def verifica_credito(self):

        contas = []
        for cell in base_contas['A']:
            contas.append(cell.value)

        if self.credito in contas:
            pass
        else:
            raise ValueError

    def edita_lancamento(cod):
        
        for cell in base_lancamentos['A']:
            if cell.value == cod:
                linha = cell.row
                data = valor_data.get()
                debito = valor_debito.get()
                credito = valor_credito.get()
                valor = valor_valor.get()
                descricao = valor_descricao_lanc.get('1.0', 'end-1c')
                Lancamento(valor, data, debito, credito, descricao)

                base_lancamentos[f'B{linha}'] = data
                base_lancamentos[f'C{linha}'] = debito
                base_lancamentos[f'D{linha}'] = credito
                base_lancamentos[f'E{linha}'] = valor
                base_lancamentos[f'F{linha}'] = descricao
                
                df.save('base.xlsx')
                atualizar_tree_lancamentos()
                limpar_campos_lancamentos()
                messagebox.showinfo(title='Sucesso!', message='Lançamento alterado com sucesso.')

    def deleta_lancamento(cod):

        for cell in base_lancamentos['A']:
            if cell.value == cod:
                base_lancamentos.delete_rows(cell.row)
                
                df.save('base.xlsx')
                atualizar_tree_lancamentos()
                limpar_campos_lancamentos()
                #messagebox.showinfo(title='Sucesso!', message='Lançamento deletado com sucesso.')


try:
    df = xl.load_workbook(r'base.xlsx')

    base_contas = df['Contas']
    base_lancamentos = df['Lançamentos']

except FileNotFoundError:
    df = xl.Workbook()
    
    df.create_sheet('Contas', 0)
    df.create_sheet('Lançamentos', 1)
    df.remove(df['Sheet'])

    base_contas = df['Contas']
    base_contas.append(['Código', 'Nome', 'Classificação', 'D/C', 'A/S'])

    base_lancamentos = df['Lançamentos']
    base_lancamentos.append(['Código', 'Data', 'Débito', 'Crédito', 'Valor', 'Descrição'])

    df.save('base.xlsx')


main = Tk()
main.title('Contabilidade')
main.geometry('1000x645')
main.iconbitmap(r'.\favicon.ico')
main.resizable(False, False)

# barra_menus = Menu(main)
# menu_arquivo = Menu(barra_menus, tearoff=0)
# menu_arquivo.add_command(label='Deletar banco de dados')
# menu_arquivo.add_separator()
# menu_arquivo.add_command(label='Sair', command=main.quit)

# menu_ajuda = Menu(barra_menus, tearoff=0)
# menu_ajuda.add_command(label='Como usar')
# menu_ajuda.add_command(label='Sobre')
# menu_ajuda.add_command(label='Encontrou um erro?')

# barra_menus.add_cascade(label='Arquivo', menu=menu_arquivo)
# barra_menus.add_cascade(label='Ajuda', menu=menu_ajuda)

# main.config(menu=barra_menus)

notebook = ttk.Notebook(main)
notebook.place(x=1, y=0, width=1000, height=625)

aba_contas = Frame(notebook)
aba_lacamentos = Frame(notebook)
aba_relatorios = Frame(notebook)

notebook.add(aba_contas, text='Contas')
notebook.add(aba_lacamentos, text='Lançamentos')
notebook.add(aba_relatorios, text='Relatorios')

cor_padrao = '#DFDFDF'
Frame(aba_contas, bg=cor_padrao).place(height=624, width=999)
Frame(aba_lacamentos, bg=cor_padrao).place(height=624, width=999)
Frame(aba_relatorios, bg=cor_padrao).place(height=624, width=999)

# ABA CONTAS #

Frame(aba_contas, bg='#000000').place(x=5, y=5, width=350, height=590)
Frame(aba_contas, bg='#D4D4D4').place(x=6, y=6, width=348, height=588)

Label(aba_contas, text='Criar contas', bg='#C0C0C0', font=('Arial', 14, 'bold')).place(x=6, y=10, width=348, height=25)

Label(aba_contas, text='Classificação:', font=('Arial', 12), bg='#D4D4D4').place(x=16, y=50, height=25)
valor_classificacao = Entry(aba_contas)
valor_classificacao.place(x=130, y=50, width=100, height=25)

Label(aba_contas, text='Descrição:', font=('Arial', 12), bg='#D4D4D4').place(x=16, y=85, height=25)
valor_descricao_conta = Entry(aba_contas)
valor_descricao_conta.place(x=130, y=85, width=100, height=25)

Label(aba_contas, text='Código:', font=('Arial', 12), bg='#D4D4D4').place(x=16, y=120, height=25)
valor_codigo = Entry(aba_contas)
valor_codigo.place(x=130, y=120, width=100, height=25)

Label(aba_contas, text='Natureza:', font=('Arial', 12), bg='#D4D4D4').place(x=16, y=155, height=25)
valor_natureza = StringVar()
valor_natureza.set('Débito')
OptionMenu(aba_contas, valor_natureza, 'Crédito', 'Débito').place(x=130, y=155, width=100, height=25)
# Label(aba_contas, text='* Somente em contas de grau 1 é necessário informar a natureza.', font=('Arial', 8), bg='#D4D4D4').place(x=16, y=190)

tree_contas = ttk.Treeview(aba_contas, columns=('Classificação', 'Descrição', 'Código'), show='headings')
tree_contas.column('Classificação', minwidth=25, width=100)
tree_contas.column('Descrição', minwidth=50, width=250)
tree_contas.column('Código', minwidth=50, width=25)
tree_contas.heading('Classificação', text='Classificação')
tree_contas.heading('Descrição', text='Descrição')
tree_contas.heading('Código', text='Código')
tree_contas.place(x=365, y=5, width=625, height=590)
tree_contas.bind('<Double-1>', duplo_clique_contas)

x_scroll = Scrollbar(tree_contas, orient=HORIZONTAL, command=tree_contas.xview)
y_scroll = Scrollbar(tree_contas, orient=VERTICAL, command=tree_contas.yview)
x_scroll.place(y=572, x=1, width=623)
y_scroll.place(y=25, x=607, height=548)
tree_contas.config(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

atualizar_tree_contas()

Button(aba_contas, text='Gravar', command=criar_conta).place(x=19, y=235, width=75, height=25)
Button(aba_contas, text='Alterar', command=editar_conta).place(x=143, y=235, width=75, height=25)
Button(aba_contas, text='Excluir', command=deletar_conta).place(x=264, y=235, width=75, height=25)

LabelFrame(aba_contas, text='Busca', bg='#D4D4D4', font=('Arial', 14)).place(x=19, y=285, width=319.5, height=150)
Label(aba_contas, text='Palavra-chave:', font=('Arial', 12), bg='#D4D4D4').place(x=23, y=315, height=25)
valor_busca_conta = Entry(aba_contas)
valor_busca_conta.place(x=26, y=345, height=25, width=305)

Button(aba_contas, text='Buscar', command=buscar_conta).place(x=76, y=390, width=75, height=25)
Button(aba_contas, text='Limpar', command=atualizar_tree_contas).place(x=206, y=390, width=75, height=25)

# ABA LANCAMENTOS #

Frame(aba_lacamentos, bg='#000000').place(x=5, y=5, width=350, height=590)
Frame(aba_lacamentos, bg='#D4D4D4').place(x=6, y=6, width=348, height=588)

tree_lancamentos = ttk.Treeview(aba_lacamentos, columns=('Id', 'Data', 'Débito', 'Crédito', 'Valor', 'Histórico'), show='headings')
tree_lancamentos.column('Id', minwidth=10, width=10)
tree_lancamentos.column('Data', minwidth=10, width=40)
tree_lancamentos.column('Débito', minwidth=10, width=20)
tree_lancamentos.column('Crédito', minwidth=10, width=20)
tree_lancamentos.column('Valor', minwidth=10, width=35)
tree_lancamentos.column('Histórico', minwidth=10, width=330)
tree_lancamentos.heading('Id', text='Id')
tree_lancamentos.heading('Data', text='Data')
tree_lancamentos.heading('Débito', text='Débito')
tree_lancamentos.heading('Crédito', text='Crédito')
tree_lancamentos.heading('Valor', text='Valor')
tree_lancamentos.heading('Histórico', text='Histórico')
tree_lancamentos.place(x=365, y=5, width=625, height=590)
tree_lancamentos.bind('<Double-1>', duplo_clique_lancamentos)

x_scroll = Scrollbar(tree_lancamentos, orient=HORIZONTAL, command=tree_lancamentos.xview)
y_scroll = Scrollbar(tree_lancamentos, orient=VERTICAL, command=tree_lancamentos.yview)
x_scroll.place(y=572, x=1, width=623)
y_scroll.place(y=25, x=607, height=548)
tree_lancamentos.config(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

atualizar_tree_lancamentos()

Label(aba_lacamentos, text='Lançar', bg='#C0C0C0', font=('Arial', 14, 'bold')).place(x=6, y=10, width=348, height=25)

Label(aba_lacamentos, text='Data:', font=('Arial', 12), bg='#D4D4D4').place(x=16, y=50, height=25)
valor_data = DateEntry(aba_lacamentos)
valor_data.place(x=100, y=50, width=100, height=25)

Label(aba_lacamentos, text='Débito:', font=('Arial', 12), bg='#D4D4D4').place(x=16, y=85, height=25)
valor_debito = Entry(aba_lacamentos)
valor_debito.place(x=100, y=85, width=100, height=25)

Label(aba_lacamentos, text='Crédito:', font=('Arial', 12), bg='#D4D4D4').place(x=16, y=120, height=25)
valor_credito = Entry(aba_lacamentos)
valor_credito.place(x=100, y=120, width=100, height=25)

Label(aba_lacamentos, text='Valor:', font=('Arial', 12), bg='#D4D4D4').place(x=16, y=155, height=25)
valor_valor = Entry(aba_lacamentos)
valor_valor.place(x=100, y=155, width=100, height=25)

Label(aba_lacamentos, text='Histórico:', font=('Arial', 12), bg='#D4D4D4').place(x=16, y=190, height=25)
valor_descricao_lanc = Text(aba_lacamentos)
valor_descricao_lanc.place(x=19, y=215, width=320, height=100)

Button(aba_lacamentos, text='Gravar', command=lancar).place(x=19, y=350, width=75, height=25)

Button(aba_lacamentos, text='Alterar', command=editar_lancamento).place(x=143, y=350, width=75, height=25)

Button(aba_lacamentos, text='Excluir', command=deletar_lancamento).place(x=264, y=350, width=75, height=25)

LabelFrame(aba_lacamentos, text='Busca', bg='#D4D4D4', font=('Arial', 14)).place(x=19, y=400, width=319.5, height=150)
Label(aba_lacamentos, text='Palavra-chave:', font=('Arial', 12), bg='#D4D4D4').place(x=23, y=430, height=25)
valor_busca_lancamento = Entry(aba_lacamentos)
valor_busca_lancamento.place(x=26, y=460, height=25, width=305)

Button(aba_lacamentos, text='Buscar', command=buscar_lancamento).place(x=76, y=505, width=75, height=25)
Button(aba_lacamentos, text='Limpar', command=atualizar_tree_lancamentos).place(x=206, y=505, width=75, height=25)

# ABA RELATORIOS #

ano = date.today().year

Frame(aba_relatorios, bg='#000000').place(x=5, y=5, width=325, height=590)
Frame(aba_relatorios, bg='#D4D4D4').place(x=6, y=6, width=323, height=588)

Frame(aba_relatorios, bg='#000000').place(x=335.5, y=5, width=325, height=590)
Frame(aba_relatorios, bg='#D4D4D4').place(x=336.5, y=6, width=323, height=588)

Frame(aba_relatorios, bg='#000000').place(x=667, y=5, width=325, height=590)
Frame(aba_relatorios, bg='#D4D4D4').place(x=668, y=6, width=323, height=588)

# razão
Label(aba_relatorios, text='Razão', bg='#C0C0C0', font=('Arial', 14, 'bold')).place(x=6, y=10, width=323, height=25)
Label(aba_relatorios, text='Conta:', font=('Arial', 12), bg='#D4D4D4').place(x=26, y=80, height=25)
valor_conta_razao = Entry(aba_relatorios)
valor_conta_razao.place(x=124, y=80, height=25, width=75)

Label(aba_relatorios, text='Período:', font=('Arial', 12), bg='#D4D4D4').place(x=26, y=125, height=25)
Label(aba_relatorios, text='a', font=('Arial', 12), bg='#D4D4D4').place(x=156.5, y=160, height=25, width=10)
valor_data_inicial_razao = DateEntry(aba_relatorios)
valor_data_inicial_razao.place(x=26, y=160, width=98, height=25)

valor_data_final_razao = DateEntry(aba_relatorios)
valor_data_final_razao.place(x=199, y=160, width=98, height=25)

valor_data_inicial_razao.delete(0, END)
valor_data_final_razao.delete(0, END)
valor_data_inicial_razao.insert(0, f'01/01/{ano}')
valor_data_final_razao.insert(0, f'31/12/{ano}')

Button(aba_relatorios, text='Gerar', command=gerar_razao).place(x=124, y=235, width=75, height=25)

# resumo
Label(aba_relatorios, text='Resumo', bg='#C0C0C0', font=('Arial', 14, 'bold')).place(x=336.5, y=10, width=323, height=25)

tree_resumo = ttk.Treeview(aba_relatorios, columns=('Mês', 'Débito', 'Crédito', 'Saldo'), show='headings')
tree_resumo.column('Mês', minwidth=10, width=10)
tree_resumo.column('Débito', minwidth=10, width=10)
tree_resumo.column('Crédito', minwidth=10, width=10)
tree_resumo.column('Saldo', minwidth=10, width=10)
tree_resumo.heading('Mês', text='Mês')
tree_resumo.heading('Débito', text='Débito')
tree_resumo.heading('Crédito', text='Crédito')
tree_resumo.heading('Saldo', text='Saldo')

tree_resumo.place(x=346, y=275, width=305, height=268)

meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
         'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

for i in meses:
    tree_resumo.insert('', END, values=(i, '0,00', '0,00', '0,00'))

Label(aba_relatorios, text='Conta:', font=('Arial', 12), bg='#D4D4D4').place(x=348, y=80, height=25, width=75)
Label(aba_relatorios, text='Ano:', font=('Arial', 12), bg='#D4D4D4').place(x=343, y=125, height=25, width=75)

valor_conta_resumo = Entry(aba_relatorios)
valor_conta_resumo.place(x=460.5, y=80, height=25, width=75)

valor_ano_resumo = Entry(aba_relatorios)
valor_ano_resumo.place(x=460.5, y=125, height=25, width=75)
valor_ano_resumo.insert(0, "2022")

Button(aba_relatorios, text='Gerar', command=gerar_resumo).place(x=460.5, y=200, width=75, height=25)

# balancete
Label(aba_relatorios, text='Balancete', bg='#C0C0C0', font=('Arial', 14, 'bold')).place(x=668, y=10, width=323, height=25)

Label(aba_relatorios, text='Período:', font=('Arial', 12), bg='#D4D4D4').place(x=699, y=80, height=25)
Label(aba_relatorios, text='a', font=('Arial', 12), bg='#D4D4D4').place(x=829.5, y=115, height=25, width=10)
valor_data_inicial_balancete = DateEntry(aba_relatorios)
valor_data_inicial_balancete.place(x=699, y=115, width=98, height=25)

valor_data_final_balancete = DateEntry(aba_relatorios)
valor_data_final_balancete.place(x=872, y=115, width=98, height=25)

valor_data_inicial_balancete.delete(0, END)
valor_data_final_balancete.delete(0, END)
valor_data_inicial_balancete.insert(0, f'01/01/{ano}')
valor_data_final_balancete.insert(0, f'31/12/{ano}')

Button(aba_relatorios, text='Gerar', command=gerar_balancete).place(x=797, y=235, width=75, height=25)

main.mainloop()
